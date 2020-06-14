import datetime
import copy
import os
import openpyxl
import openpyxl.utils
import logging
import numpy as np
from modules.utilities import toolbox

class InvoicesFromTemplate: # todo refactor into multiple classes

    def __init__(self, times, clients, invoices, params):
        self.times = times
        self.clients = clients
        self.invoices = invoices
        self.params = params

        self.logger = logging.getLogger('main.' + __name__)

        try:
            self.template_workbook = openpyxl.load_workbook(params.invoice_template_path)
            self.worksheet_name = self.template_workbook.sheetnames[0]
            self.template_worksheet = self.template_workbook.get_sheet_by_name(self.worksheet_name)
            self._generate_cell_addresses_from_template_dict()
            self._determine_working_day_rows()
        except FileNotFoundError:
            self.logger.error('Template file was not found at {}. No invoices can be generated without one. Please make a template file.'.format(params.invoice_template_path))

    def _generate_cell_addresses_from_template_dict(self):
        cell_coordinate_attribute_get_info_structure_df_itertuples_method_tuple_list = [('times_cell_coordinate_dict', 'get_times_info_structure_df_itertuples'),
                                                                                    ('clients_cell_coordinate_dict', 'get_clients_info_structure_df_itertuples'),
                                                                                    ('invoices_cell_coordinate_dict', 'get_invoices_info_structure_df_itertuples'),
                                                                                    ('calculated_cell_coordinate_dict', 'get_calculations_info_structure_df_itertuples')]

        for cell_coordinate_attribute, get_info_structure_df_itertuples_method in cell_coordinate_attribute_get_info_structure_df_itertuples_method_tuple_list:
            self._generate_cell_coordinate_dict_from_info_structure_df(cell_coordinate_attribute, get_info_structure_df_itertuples_method)

    def _generate_cell_coordinate_dict_from_info_structure_df(self, cell_coordinate_attribute, get_info_structure_df_itertuples_method):
        setattr(self, cell_coordinate_attribute, dict())
        for info_structure_row in getattr(self.params, get_info_structure_df_itertuples_method)():
            info_name = info_structure_row.info_name
            invoice_template_tag = info_structure_row.invoice_template_tag
            if isinstance(invoice_template_tag, str):
                for row in range(1, self.template_worksheet.max_row + 1):
                    for col in range(1, self.template_worksheet.max_column + 1):
                        cell = self.template_worksheet.cell(row=row, column=col)
                        if cell.value == invoice_template_tag:
                            getattr(self, cell_coordinate_attribute)[info_name] = cell.coordinate

    def _determine_working_day_rows(self):
        self.working_day_start_row, _ = openpyxl.utils.cell.coordinate_to_tuple(self.times_cell_coordinate_dict['Type of hours'])
        self.working_day_stop_row, _ = openpyxl.utils.cell.coordinate_to_tuple(self.calculated_cell_coordinate_dict['Calculated amount for working day'])
        self.row_increment_between_two_working_days = self.working_day_stop_row - self.working_day_start_row + 2

    def generate_invoice(self, invoice_info_dict):
        invoice_workbook = self._initialize_workbook()
        client_info_dict, invoice_times_df = self._create_client_info_dict_and_invoice_times_df(invoice_info_dict)

        self._copy_and_fill_in_workbook_begin(client_info_dict, invoice_info_dict)
        self._copy_and_fill_in_working_days(invoice_times_df)
        self._copy_and_fill_in_workbook_end(invoice_times_df)

        self._save_invoice_workbook(invoice_workbook, invoice_info_dict)

    def _initialize_workbook(self):
        invoice_workbook = openpyxl.Workbook()
        blank_worksheet_name = invoice_workbook.sheetnames[0]
        blank_worksheet = invoice_workbook.get_sheet_by_name(blank_worksheet_name)
        invoice_workbook.remove_sheet(blank_worksheet)
        self.invoice_worksheet = invoice_workbook.create_sheet(self.worksheet_name)  # todo get somewhere else
        return invoice_workbook

    def _create_client_info_dict_and_invoice_times_df(self, invoice_info_dict):
        client_name = invoice_info_dict['Client name']
        client_info_dict = self.clients.get_client_info_dict_based_on_client_name(client_name)

        invoice_month = invoice_info_dict['Month']
        invoice_year = invoice_info_dict['Year']

        start_date, stop_date = self._extract_start_and_stop_date_from_month_and_year(invoice_month, invoice_year)
        full_times_df = self.times.get_times_df()

        valid_row_indices = (start_date <= full_times_df['Date']) & (full_times_df['Date'] < stop_date) & (full_times_df['Client name'] == client_name)
        invoice_times_df = full_times_df.loc[valid_row_indices]
        invoice_times_df = invoice_times_df.sort_values(by='Date', ignore_index=True)
        return client_info_dict, invoice_times_df

    def _extract_start_and_stop_date_from_month_and_year(self, month, year):
        date_string = month + ' ' + str(year)
        start_date = datetime.datetime.strptime(date_string, '%B %Y')

        stop_date = datetime.datetime.strptime(date_string, '%B %Y')
        stop_date_month = stop_date.month
        stop_date_year = stop_date.year
        if stop_date_month == 12:
            stop_date_month = 1
            stop_date_year = stop_date_year + 1
        else:
            stop_date_month = stop_date_month + 1
        stop_date = datetime.datetime(year=stop_date_year, month=stop_date_month, day=1, hour=0, minute=0, second=0)
        return start_date, stop_date


    def _copy_and_fill_in_workbook_begin(self, client_info_dict, invoice_info_dict):
        self._copy_workbook_begin()
        self._fill_in_invoice_info(invoice_info_dict)
        self._fill_in_client_info(client_info_dict)

    def _copy_workbook_begin(self):
        for df_row in range(1, self.working_day_start_row):
            for col in range(1, self.template_worksheet.max_column + 1):
                template_cell = self.template_worksheet.cell(row=df_row, column=col)
                invoice_cell = self.invoice_worksheet.cell(row=df_row, column=col)
                InvoicesFromTemplate.copy_cell(template_cell, invoice_cell)

    def _fill_in_invoice_info(self, invoice_info_dict):
        for key, val in invoice_info_dict.items():
            try:
                coordinate = self.invoices_cell_coordinate_dict[key]
                template_cell = self.invoice_worksheet[coordinate]
                template_cell.value = val
            except KeyError:
                pass

    def _fill_in_client_info(self, client_info_dict):
        for key, val in client_info_dict.items():
            try:
                coordinate = self.clients_cell_coordinate_dict[key]
                template_cell = self.invoice_worksheet[coordinate]
                template_cell.value = val
            except KeyError:
                pass

    def _copy_and_fill_in_working_days(self, invoice_times_df):
        for index, df_row in invoice_times_df.iterrows():
            row_offset = index * self.row_increment_between_two_working_days
            self._copy_working_day(row_offset)
            self._fill_in_working_day(df_row, row_offset)
            self._fill_in_calculated_hours(row_offset)

            if df_row['Type of hours'] == 'during day':
                self._fill_in_calculated_amount_for_hours_during_day(row_offset)
            elif df_row['Type of hours'] == 'shift':
                self._fill_in_calculated_amount_for_hours_in_shift(row_offset)
            else:
                self.logger.error('Invalid type of hours in times')

            self._fill_in_calculated_amount_for_commute(row_offset)
            self._fill_in_calculated_amount_for_distance_during_work(row_offset)
            self._fill_in_calculated_amount_for_working_day(row_offset)

    def _copy_working_day(self, row_offset):
        for row in range(self.working_day_start_row, self.working_day_stop_row + 1):
            for col in range(1, self.template_worksheet.max_column + 1):
                template_cell = self.template_worksheet.cell(row=row, column=col)
                invoice_row = row + row_offset
                invoice_cell = self.invoice_worksheet.cell(row=invoice_row, column=col)
                InvoicesFromTemplate.copy_cell(template_cell, invoice_cell)

    def _fill_in_working_day(self, df_row, row_offset): # todo clean up
        for key, val in df_row.iteritems():
            try:
                if key == 'Type of hours':
                    if val == 'during day':
                        self.set_times_value(key, row_offset, self.params.translation_during_day)
                    elif val == 'shift':
                        self.set_times_value(key, row_offset, self.params.translation_shift)
                    else:
                        self.logger.error('Invalid type of hours in times')
                else:
                    self.set_times_value(key, row_offset, val)
            except KeyError:
                pass

    def set_times_value(self, info_name, row_offset, value):
        coordinate = self.times_cell_coordinate_dict[info_name]
        initial_row, initial_col = openpyxl.utils.cell.coordinate_to_tuple(coordinate)
        used_row = initial_row + row_offset
        cell = self.invoice_worksheet.cell(row=used_row, column=initial_col)
        cell.value = value

    def _fill_in_calculated_hours(self, row_offset):
        initial_start_time_row, start_time_col = openpyxl.utils.cell.coordinate_to_tuple(self.times_cell_coordinate_dict['Start time'])
        initial_stop_time_row, stop_time_col = openpyxl.utils.cell.coordinate_to_tuple(self.times_cell_coordinate_dict['Stop time'])

        start_time_row = initial_start_time_row + row_offset
        stop_time_row = initial_stop_time_row + row_offset

        start_time_coordinate = InvoicesFromTemplate.tuple_to_coordinate(start_time_row, start_time_col)
        stop_time_coordinate = InvoicesFromTemplate.tuple_to_coordinate(stop_time_row, stop_time_col)

        self.set_calculated_value('Calculated hours', row_offset, '=24*({}-{})', [stop_time_coordinate, start_time_coordinate])

    def _fill_in_calculated_amount_for_hours_during_day(self, row_offset):
        multiplication_coordinate = self.invoices_cell_coordinate_dict['Rate during day (euro/h)']
        quantity_coordinate = self.calculated_cell_coordinate_dict['Calculated hours']
        self._fill_in_calculated_amount('Calculated amount for hours', row_offset, multiplication_coordinate, quantity_coordinate)

    def _fill_in_calculated_amount_for_hours_in_shift(self, row_offset):
        multiplication_coordinate = self.invoices_cell_coordinate_dict['Rate for shifts (euro/h)']
        quantity_coordinate = self.calculated_cell_coordinate_dict['Calculated hours']
        self._fill_in_calculated_amount('Calculated amount for hours', row_offset, multiplication_coordinate, quantity_coordinate)

    def _fill_in_calculated_amount_for_commute(self, row_offset):
        multiplication_coordinate = self.invoices_cell_coordinate_dict['Compensation for commute (euro/km)']
        quantity_coordinate = self.times_cell_coordinate_dict['Commute (km)']
        self._fill_in_calculated_amount('Calculated amount for commute', row_offset, multiplication_coordinate, quantity_coordinate)

    def _fill_in_calculated_amount_for_distance_during_work(self, row_offset):
        multiplication_coordinate = self.invoices_cell_coordinate_dict['Compensation for driving during work (euro/km)']
        quantity_coordinate = self.times_cell_coordinate_dict['Distance during work (km)']
        self._fill_in_calculated_amount('Calculated amount for distance during work', row_offset, multiplication_coordinate, quantity_coordinate)

    def _fill_in_calculated_amount(self, calculated_amount, row_offset, multiplication_coordinate, quantity_coordinate):
        new_multiplication_coordinate = openpyxl.utils.cell.absolute_coordinate(multiplication_coordinate)
        new_quantity_coordinate = InvoicesFromTemplate._increment_coordinate_by_row_offset(quantity_coordinate, row_offset)
        self.set_calculated_value(calculated_amount, row_offset, '={}*{}', [new_multiplication_coordinate, new_quantity_coordinate])

    def _fill_in_calculated_amount_for_working_day(self, row_offset):
        coordinate_list = [self.calculated_cell_coordinate_dict['Calculated amount for hours'], self.calculated_cell_coordinate_dict['Calculated amount for commute'], self.calculated_cell_coordinate_dict['Calculated amount for distance during work']]
        new_coordinate_list = [InvoicesFromTemplate._increment_coordinate_by_row_offset(coordinate, row_offset) for coordinate in coordinate_list]
        string_format = '=' + '+'.join(['{}']*len(coordinate_list)) # should give e.g. '={}+{}' if there are two coordinates
        self.set_calculated_value('Calculated amount for working day', row_offset, string_format, new_coordinate_list)

    def _copy_and_fill_in_workbook_end(self, invoice_times_df):
        row_offset_for_end = (len(invoice_times_df) - 1) * self.row_increment_between_two_working_days
        self._copy_workbook_end(row_offset_for_end)
        self._fill_in_calculated_total_amount(invoice_times_df, row_offset_for_end)

    def _copy_workbook_end(self, row_offset_for_end):
        for row in range(self.working_day_stop_row + 2, self.template_worksheet.max_row + 1):
            for col in range(1, self.template_worksheet.max_column + 1):
                template_cell = self.template_worksheet.cell(row=row, column=col)
                invoice_row = row + row_offset_for_end
                invoice_cell = self.invoice_worksheet.cell(row=invoice_row, column=col)
                InvoicesFromTemplate.copy_cell(template_cell, invoice_cell)

    def _fill_in_calculated_total_amount(self, invoice_times_df, row_offset):
        number_of_working_days = len(invoice_times_df)
        initial_calculated_amount_for_working_day_coordinate = self.calculated_cell_coordinate_dict['Calculated amount for working day']
        new_coordinate_list = [InvoicesFromTemplate._increment_coordinate_by_row_offset(initial_calculated_amount_for_working_day_coordinate, index*self.row_increment_between_two_working_days) for index in range(number_of_working_days)]
        string_format = '=' + '+'.join(['{}']*number_of_working_days) # should give e.g. '={}+{}' if there are two working days
        self.set_calculated_value('Calculated total amount', row_offset, string_format, new_coordinate_list)

    def _save_invoice_workbook(self, invoice_workbook, invoice_dict):
        invoice_path = invoice_dict['File path']
        invoice_workbook.save(invoice_path)
        self.logger.info('Invoice saved to location: {}'.format(invoice_path))

    def set_calculated_value(self, info_name, row_offset, string_format, string_parameter_list):
        coordinate = self.calculated_cell_coordinate_dict[info_name]
        initial_row, initial_col = openpyxl.utils.cell.coordinate_to_tuple(coordinate)
        used_row = initial_row + row_offset
        cell = self.invoice_worksheet.cell(row=used_row, column=initial_col)
        value = string_format.format(*string_parameter_list)
        cell.value = value

    @staticmethod
    def _increment_coordinate_by_row_offset(initial_coordinate, row_offset):
        initial_row, col = openpyxl.utils.cell.coordinate_to_tuple(initial_coordinate)
        new_row = initial_row + row_offset
        new_coordinate = InvoicesFromTemplate.tuple_to_coordinate(new_row, col)
        return new_coordinate

    @staticmethod
    def copy_cell(cell, new_cell):
        new_cell.value = cell.value
        if cell.has_style:
            new_cell.font = copy.copy(cell.font)
            new_cell.border = copy.copy(cell.border)
            new_cell.fill = copy.copy(cell.fill)
            new_cell.number_format = copy.copy(cell.number_format)
            new_cell.protection = copy.copy(cell.protection)
            new_cell.alignment = copy.copy(cell.alignment)

    @staticmethod
    def tuple_to_coordinate(row, col):
        col_letter = openpyxl.utils.cell.get_column_letter(col)
        coordinate = '{}{}'.format(col_letter, row)
        return coordinate

